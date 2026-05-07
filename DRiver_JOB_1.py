"""
Delivery Driver Profit Calculator UK 2026
==========================================
A professional GUI application for UK courier/van drivers to calculate
daily, weekly, monthly and annual net profit with tax estimation.

Author: Generated for UK Delivery Drivers
Python: 3.13+
Dependencies: PySide6, matplotlib, openpyxl, pandas
"""

import sys
import json
import csv
import os
from pathlib import Path
from datetime import datetime

# ---------------------------------------------------------------------------
# Dependency check & graceful error
# ---------------------------------------------------------------------------
try:
    from PySide6.QtWidgets import (
        QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
        QGridLayout, QTabWidget, QLabel, QLineEdit, QComboBox, QPushButton,
        QScrollArea, QFrame, QGroupBox, QSplitter, QFileDialog, QMessageBox,
        QTableWidget, QTableWidgetItem, QHeaderView, QDialog, QFormLayout,
        QSpinBox, QDoubleSpinBox, QTextEdit, QListWidget, QListWidgetItem,
        QSizePolicy, QProgressBar
    )
    from PySide6.QtCore import Qt, QTimer, Signal, QThread, QSize
    from PySide6.QtGui import (
        QFont, QColor, QPalette, QIcon, QPixmap, QPainter,
        QLinearGradient, QBrush, QFontDatabase
    )
except ImportError:
    print("ERROR: PySide6 not found. Install with: pip install PySide6")
    sys.exit(1)

try:
    import matplotlib
    matplotlib.use("QtAgg")
    from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
    from matplotlib.figure import Figure
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    HAS_MATPLOTLIB = True
except ImportError:
    HAS_MATPLOTLIB = False
    print("WARNING: matplotlib not found. Charts disabled. Install: pip install matplotlib")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

# ---------------------------------------------------------------------------
# UK 2025/26 Tax Constants
# ---------------------------------------------------------------------------
UK_TAX = {
    "personal_allowance": 12_570,
    "basic_rate_limit": 50_270,
    "higher_rate_limit": 125_140,
    "basic_rate": 0.20,
    "higher_rate": 0.40,
    "additional_rate": 0.45,
    # NI Employee (Class 1)
    "ni_primary_threshold": 12_570,
    "ni_upper_earnings": 50_270,
    "ni_basic_rate": 0.08,
    "ni_higher_rate": 0.02,
    # NI Self Employed (Class 4)
    "ni_se_lower": 12_570,
    "ni_se_upper": 50_270,
    "ni_se_basic": 0.09,
    "ni_se_higher": 0.02,
    # Class 2 NI (self-employed flat weekly)
    "ni_class2_weekly": 3.45,
    "ni_class2_threshold": 12_570,
    # Minimum wage 2026 estimate (21+)
    "minimum_wage": 12.21,
}

# ---------------------------------------------------------------------------
# Data Models
# ---------------------------------------------------------------------------

class JobProfile:
    """Stores all input data for a single delivery job profile."""

    DEFAULT_NAME = "New Job"

    def __init__(self, name: str = DEFAULT_NAME):
        self.name = name
        self.created = datetime.now().isoformat()
        # Income
        self.day_rate: float = 120.0
        self.days_per_week: int = 5
        self.hours_per_day: float = 9.0
        self.mileage_bonus_weekly: float = 0.0
        self.performance_bonus_weekly: float = 0.0
        self.training_payment: float = 0.0
        # Vehicle expenses (weekly)
        self.van_rental: float = 0.0
        self.fuel: float = 0.0
        self.insurance: float = 0.0
        self.parking_fines: float = 0.0
        self.repairs: float = 0.0
        # Other expenses (weekly)
        self.phone_data: float = 0.0
        self.uniform_equipment: float = 0.0
        self.other_costs: float = 0.0
        # Tax
        self.employment_type: str = "Self Employed"  # or "PAYE Employee"

    def to_dict(self) -> dict:
        return self.__dict__.copy()

    @classmethod
    def from_dict(cls, data: dict) -> "JobProfile":
        p = cls(data.get("name", cls.DEFAULT_NAME))
        for k, v in data.items():
            if hasattr(p, k):
                setattr(p, k, v)
        return p


class TaxCalculator:
    """UK 2025/26 Income Tax and National Insurance calculations."""

    @staticmethod
    def calculate_income_tax(gross_annual: float) -> float:
        """Calculate annual income tax for a given gross annual income."""
        taxable = max(0.0, gross_annual - UK_TAX["personal_allowance"])
        if taxable <= 0:
            return 0.0
        basic_band = min(taxable, UK_TAX["basic_rate_limit"] - UK_TAX["personal_allowance"])
        tax = basic_band * UK_TAX["basic_rate"]
        if taxable > (UK_TAX["basic_rate_limit"] - UK_TAX["personal_allowance"]):
            higher_band = min(
                taxable - (UK_TAX["basic_rate_limit"] - UK_TAX["personal_allowance"]),
                UK_TAX["higher_rate_limit"] - UK_TAX["basic_rate_limit"]
            )
            tax += higher_band * UK_TAX["higher_rate"]
        if taxable > (UK_TAX["higher_rate_limit"] - UK_TAX["personal_allowance"]):
            additional_band = taxable - (UK_TAX["higher_rate_limit"] - UK_TAX["personal_allowance"])
            tax += additional_band * UK_TAX["additional_rate"]
        return tax

    @staticmethod
    def calculate_ni_employee(gross_annual: float) -> float:
        """Class 1 NI for PAYE employees."""
        if gross_annual <= UK_TAX["ni_primary_threshold"]:
            return 0.0
        basic = min(gross_annual, UK_TAX["ni_upper_earnings"]) - UK_TAX["ni_primary_threshold"]
        ni = basic * UK_TAX["ni_basic_rate"]
        if gross_annual > UK_TAX["ni_upper_earnings"]:
            ni += (gross_annual - UK_TAX["ni_upper_earnings"]) * UK_TAX["ni_higher_rate"]
        return ni

    @staticmethod
    def calculate_ni_self_employed(gross_annual: float) -> float:
        """Class 2 + Class 4 NI for self-employed."""
        class2 = 0.0
        if gross_annual >= UK_TAX["ni_class2_threshold"]:
            class2 = UK_TAX["ni_class2_weekly"] * 52

        class4 = 0.0
        if gross_annual > UK_TAX["ni_se_lower"]:
            basic = min(gross_annual, UK_TAX["ni_se_upper"]) - UK_TAX["ni_se_lower"]
            class4 = basic * UK_TAX["ni_se_basic"]
            if gross_annual > UK_TAX["ni_se_upper"]:
                class4 += (gross_annual - UK_TAX["ni_se_upper"]) * UK_TAX["ni_se_higher"]
        return class2 + class4


class ProfitCalculator:
    """Core profit calculation engine."""

    def __init__(self, profile: JobProfile):
        self.p = profile
        self._results: dict = {}

    def calculate(self) -> dict:
        p = self.p
        weeks_per_month = 52 / 12  # ~4.333

        # --- Gross Income ---
        gross_daily = p.day_rate
        gross_weekly = (p.day_rate * p.days_per_week) + p.mileage_bonus_weekly + p.performance_bonus_weekly
        gross_monthly = gross_weekly * weeks_per_month
        gross_annual = gross_weekly * 52 + p.training_payment

        # --- Expenses ---
        weekly_expenses = (
            p.van_rental + p.fuel + p.insurance +
            p.parking_fines + p.repairs +
            p.phone_data + p.uniform_equipment + p.other_costs
        )
        monthly_expenses = weekly_expenses * weeks_per_month
        annual_expenses = weekly_expenses * 52

        # Expense breakdown categories
        vehicle_costs_weekly = p.van_rental + p.fuel + p.insurance + p.parking_fines + p.repairs
        other_costs_weekly = p.phone_data + p.uniform_equipment + p.other_costs

        # --- Tax ---
        net_annual_before_tax = gross_annual - annual_expenses
        if p.employment_type == "PAYE Employee":
            income_tax_annual = TaxCalculator.calculate_income_tax(gross_annual)
            ni_annual = TaxCalculator.calculate_ni_employee(gross_annual)
        else:
            taxable_profit = max(0.0, net_annual_before_tax)
            income_tax_annual = TaxCalculator.calculate_income_tax(taxable_profit)
            ni_annual = TaxCalculator.calculate_ni_self_employed(taxable_profit)

        total_tax_annual = income_tax_annual + ni_annual
        tax_weekly = total_tax_annual / 52
        tax_monthly = total_tax_annual / 12
        tax_reserve_pct = (total_tax_annual / gross_annual * 100) if gross_annual > 0 else 0

        # --- Net Profit ---
        net_weekly = gross_weekly - weekly_expenses - tax_weekly
        net_monthly = gross_monthly - monthly_expenses - tax_monthly
        net_annual = gross_annual - annual_expenses - total_tax_annual
        net_daily = net_weekly / p.days_per_week if p.days_per_week > 0 else 0

        # --- Effective hourly rate ---
        weekly_hours = p.hours_per_day * p.days_per_week
        hourly_rate_gross = (gross_weekly / weekly_hours) if weekly_hours > 0 else 0
        hourly_rate_net = (net_weekly / weekly_hours) if weekly_hours > 0 else 0

        # --- Warnings ---
        warnings = []
        if hourly_rate_net < UK_TAX["minimum_wage"]:
            warnings.append(
                f"⚠ LOW PROFITABILITY: Net hourly rate £{hourly_rate_net:.2f} is below "
                f"UK minimum wage £{UK_TAX['minimum_wage']:.2f}/hr"
            )
        if gross_weekly > 0 and (vehicle_costs_weekly / gross_weekly) > 0.35:
            pct = vehicle_costs_weekly / gross_weekly * 100
            warnings.append(
                f"⚠ HIGH VEHICLE COSTS: Vehicle expenses are {pct:.1f}% of gross income (threshold: 35%)"
            )
        if net_annual < 0:
            warnings.append("⚠ RUNNING AT A LOSS: Annual expenses + tax exceed gross income")

        self._results = {
            # Daily
            "gross_daily": gross_daily,
            "net_daily": net_daily,
            "hourly_rate_gross": hourly_rate_gross,
            "hourly_rate_net": hourly_rate_net,
            # Weekly
            "gross_weekly": gross_weekly,
            "weekly_expenses": weekly_expenses,
            "tax_weekly": tax_weekly,
            "net_weekly": net_weekly,
            "vehicle_costs_weekly": vehicle_costs_weekly,
            "other_costs_weekly": other_costs_weekly,
            # Monthly
            "gross_monthly": gross_monthly,
            "monthly_expenses": monthly_expenses,
            "tax_monthly": tax_monthly,
            "net_monthly": net_monthly,
            # Annual
            "gross_annual": gross_annual,
            "annual_expenses": annual_expenses,
            "income_tax_annual": income_tax_annual,
            "ni_annual": ni_annual,
            "total_tax_annual": total_tax_annual,
            "net_annual": net_annual,
            # Meta
            "tax_reserve_pct": tax_reserve_pct,
            "weekly_hours": weekly_hours,
            "warnings": warnings,
            "employment_type": p.employment_type,
        }
        return self._results

    def get_scenario(self, scenario: str) -> dict:
        """Return modified results for Good/Realistic/Worst scenarios."""
        import copy
        original = copy.deepcopy(self.p)
        p = self.p

        if scenario == "Good":
            p.day_rate = original.day_rate * 1.15
            p.mileage_bonus_weekly = original.mileage_bonus_weekly * 1.3 + 20
            p.performance_bonus_weekly = original.performance_bonus_weekly * 1.5 + 15
            p.fuel = original.fuel * 0.85
            p.parking_fines = 0
        elif scenario == "Realistic":
            pass  # use current values
        elif scenario == "Worst":
            p.day_rate = original.day_rate * 0.85
            p.mileage_bonus_weekly = 0
            p.performance_bonus_weekly = 0
            p.fuel = original.fuel * 1.25
            p.parking_fines = original.parking_fines + 15
            p.repairs = original.repairs * 2 + 20

        result = self.calculate()
        # Restore
        for k, v in original.__dict__.items():
            setattr(p, k, v)
        return result


# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

STYLE_SHEET = """
QMainWindow {
    background-color: #0f1117;
}
QWidget {
    background-color: #0f1117;
    color: #e8eaf0;
    font-family: 'Segoe UI', 'SF Pro Display', Arial, sans-serif;
    font-size: 13px;
}
QTabWidget::pane {
    border: 1px solid #2a2d3a;
    background-color: #13161f;
    border-radius: 8px;
}
QTabBar::tab {
    background-color: #1a1d28;
    color: #7b8094;
    padding: 10px 22px;
    border: none;
    border-top-left-radius: 6px;
    border-top-right-radius: 6px;
    font-weight: 600;
    font-size: 12px;
    letter-spacing: 0.5px;
}
QTabBar::tab:selected {
    background-color: #1e6fff;
    color: #ffffff;
}
QTabBar::tab:hover:!selected {
    background-color: #252836;
    color: #c0c4d4;
}
QGroupBox {
    border: 1px solid #252836;
    border-radius: 8px;
    margin-top: 14px;
    padding: 12px 8px 8px 8px;
    background-color: #13161f;
    font-weight: 700;
    font-size: 12px;
    color: #7b8094;
    letter-spacing: 0.8px;
}
QGroupBox::title {
    subcontrol-origin: margin;
    subcontrol-position: top left;
    left: 12px;
    top: 2px;
    padding: 0 6px;
    background-color: #13161f;
    color: #1e6fff;
    font-size: 11px;
    letter-spacing: 1px;
}
QLineEdit, QDoubleSpinBox, QSpinBox, QComboBox {
    background-color: #1a1d28;
    border: 1px solid #2a2d3a;
    border-radius: 6px;
    padding: 7px 10px;
    color: #e8eaf0;
    font-size: 13px;
    selection-background-color: #1e6fff;
}
QLineEdit:focus, QDoubleSpinBox:focus, QSpinBox:focus, QComboBox:focus {
    border: 1px solid #1e6fff;
    background-color: #1d2030;
}
QDoubleSpinBox::up-button, QDoubleSpinBox::down-button,
QSpinBox::up-button, QSpinBox::down-button {
    background-color: #252836;
    border: none;
    width: 18px;
}
QDoubleSpinBox::up-button:hover, QDoubleSpinBox::down-button:hover,
QSpinBox::up-button:hover, QSpinBox::down-button:hover {
    background-color: #1e6fff;
}
QComboBox::drop-down {
    border: none;
    background-color: #252836;
    width: 24px;
    border-radius: 0 6px 6px 0;
}
QComboBox QAbstractItemView {
    background-color: #1a1d28;
    border: 1px solid #2a2d3a;
    selection-background-color: #1e6fff;
    color: #e8eaf0;
}
QPushButton {
    background-color: #1e6fff;
    color: #ffffff;
    border: none;
    border-radius: 7px;
    padding: 9px 20px;
    font-weight: 700;
    font-size: 12px;
    letter-spacing: 0.5px;
}
QPushButton:hover {
    background-color: #3d7fff;
}
QPushButton:pressed {
    background-color: #1558cc;
}
QPushButton.secondary {
    background-color: #252836;
    color: #c0c4d4;
}
QPushButton.secondary:hover {
    background-color: #2e3245;
}
QPushButton.danger {
    background-color: #c0392b;
}
QPushButton.success {
    background-color: #27ae60;
}
QLabel.section-header {
    font-size: 16px;
    font-weight: 800;
    color: #ffffff;
    letter-spacing: 0.3px;
}
QLabel.metric-value {
    font-size: 22px;
    font-weight: 800;
    color: #1e6fff;
}
QLabel.metric-label {
    font-size: 11px;
    color: #7b8094;
    letter-spacing: 0.8px;
}
QLabel.warning-label {
    color: #f39c12;
    font-size: 12px;
    font-weight: 600;
    padding: 6px 10px;
    background-color: #2d2408;
    border-left: 3px solid #f39c12;
    border-radius: 4px;
}
QLabel.positive {
    color: #27ae60;
}
QLabel.negative {
    color: #e74c3c;
}
QScrollArea {
    border: none;
    background-color: transparent;
}
QScrollBar:vertical {
    background-color: #1a1d28;
    width: 8px;
    border-radius: 4px;
}
QScrollBar::handle:vertical {
    background-color: #3a3d52;
    border-radius: 4px;
    min-height: 20px;
}
QScrollBar::handle:vertical:hover {
    background-color: #1e6fff;
}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
    height: 0px;
}
QListWidget {
    background-color: #1a1d28;
    border: 1px solid #2a2d3a;
    border-radius: 6px;
    padding: 4px;
}
QListWidget::item {
    padding: 8px 10px;
    border-radius: 4px;
    color: #c0c4d4;
}
QListWidget::item:selected {
    background-color: #1e6fff;
    color: #ffffff;
}
QListWidget::item:hover:!selected {
    background-color: #252836;
}
QTableWidget {
    background-color: #13161f;
    border: 1px solid #2a2d3a;
    border-radius: 6px;
    gridline-color: #1e2130;
    color: #e8eaf0;
}
QTableWidget::item {
    padding: 6px 10px;
}
QTableWidget::item:selected {
    background-color: #1e3a6e;
}
QHeaderView::section {
    background-color: #1a1d28;
    color: #7b8094;
    border: none;
    border-right: 1px solid #252836;
    padding: 8px 10px;
    font-weight: 700;
    font-size: 11px;
    letter-spacing: 0.8px;
}
QTextEdit {
    background-color: #1a1d28;
    border: 1px solid #2a2d3a;
    border-radius: 6px;
    color: #e8eaf0;
    padding: 8px;
}
QSplitter::handle {
    background-color: #2a2d3a;
    width: 2px;
}
"""


# ---------------------------------------------------------------------------
# Reusable Widgets
# ---------------------------------------------------------------------------

class MoneySpinBox(QDoubleSpinBox):
    """A styled spin box for monetary values."""
    def __init__(self, parent=None, max_val=9999.99):
        super().__init__(parent)
        self.setPrefix("£ ")
        self.setRange(0, max_val)
        self.setDecimals(2)
        self.setSingleStep(1.0)
        self.setFixedHeight(34)

    def textFromValue(self, val):
        return f"{val:,.2f}"


class MetricCard(QFrame):
    """A card widget displaying a single metric."""
    def __init__(self, label: str, value: str = "—", color: str = "#1e6fff", parent=None):
        super().__init__(parent)
        self.setFrameShape(QFrame.StyledPanel)
        self.setStyleSheet(f"""
            QFrame {{
                background-color: #1a1d28;
                border: 1px solid #252836;
                border-radius: 10px;
                border-top: 3px solid {color};
            }}
        """)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(14, 12, 14, 12)
        layout.setSpacing(4)

        self.label_widget = QLabel(label.upper())
        self.label_widget.setStyleSheet("color: #7b8094; font-size: 10px; font-weight: 700; letter-spacing: 1.2px;")

        self.value_widget = QLabel(value)
        self.value_widget.setStyleSheet(f"color: {color}; font-size: 20px; font-weight: 800;")

        layout.addWidget(self.label_widget)
        layout.addWidget(self.value_widget)

    def update_value(self, value: str):
        self.value_widget.setText(value)

    def set_color(self, color: str):
        self.value_widget.setStyleSheet(f"color: {color}; font-size: 20px; font-weight: 800;")


class SectionHeader(QLabel):
    def __init__(self, text: str, parent=None):
        super().__init__(text, parent)
        self.setStyleSheet("""
            font-size: 14px;
            font-weight: 800;
            color: #ffffff;
            padding: 4px 0;
            letter-spacing: 0.3px;
        """)


class InputRow(QWidget):
    """Label + input widget in a horizontal layout."""
    def __init__(self, label: str, widget: QWidget, parent=None):
        super().__init__(parent)
        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 2, 0, 2)
        layout.setSpacing(10)

        lbl = QLabel(label)
        lbl.setFixedWidth(200)
        lbl.setStyleSheet("color: #a0a4b8; font-size: 12px;")
        lbl.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)

        layout.addWidget(lbl)
        layout.addWidget(widget)
        layout.addStretch()


# ---------------------------------------------------------------------------
# Tab: Inputs
# ---------------------------------------------------------------------------

class InputsTab(QScrollArea):
    calculation_requested = Signal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWidgetResizable(True)

        container = QWidget()
        self.setWidget(container)
        main_layout = QVBoxLayout(container)
        main_layout.setSpacing(14)
        main_layout.setContentsMargins(16, 16, 16, 16)

        # --- Income Group ---
        income_group = QGroupBox("INCOME")
        ig_layout = QVBoxLayout(income_group)
        ig_layout.setSpacing(6)

        self.day_rate = MoneySpinBox(max_val=999.99)
        self.day_rate.setValue(120.0)

        self.days_per_week = QSpinBox()
        self.days_per_week.setRange(1, 7)
        self.days_per_week.setValue(5)
        self.days_per_week.setFixedHeight(34)
        self.days_per_week.setSuffix(" days")

        self.hours_per_day = QDoubleSpinBox()
        self.hours_per_day.setRange(1, 24)
        self.hours_per_day.setValue(9.0)
        self.hours_per_day.setSuffix(" hrs")
        self.hours_per_day.setFixedHeight(34)

        self.mileage_bonus = MoneySpinBox(max_val=999.99)
        self.performance_bonus = MoneySpinBox(max_val=999.99)
        self.training_payment = MoneySpinBox(max_val=9999.99)

        ig_layout.addWidget(InputRow("Day Rate", self.day_rate))
        ig_layout.addWidget(InputRow("Days per Week", self.days_per_week))
        ig_layout.addWidget(InputRow("Hours per Day", self.hours_per_day))
        ig_layout.addWidget(InputRow("Mileage Bonus (weekly)", self.mileage_bonus))
        ig_layout.addWidget(InputRow("Performance Bonus (weekly)", self.performance_bonus))
        ig_layout.addWidget(InputRow("Training / Onboarding (one-off)", self.training_payment))

        # --- Vehicle Expenses Group ---
        vehicle_group = QGroupBox("VEHICLE EXPENSES (WEEKLY)")
        vg_layout = QVBoxLayout(vehicle_group)
        vg_layout.setSpacing(6)

        self.van_rental = MoneySpinBox(max_val=999.99)
        self.fuel = MoneySpinBox(max_val=999.99)
        self.insurance = MoneySpinBox(max_val=999.99)
        self.parking_fines = MoneySpinBox(max_val=999.99)
        self.repairs = MoneySpinBox(max_val=999.99)

        vg_layout.addWidget(InputRow("Van Rental", self.van_rental))
        vg_layout.addWidget(InputRow("Fuel", self.fuel))
        vg_layout.addWidget(InputRow("Insurance", self.insurance))
        vg_layout.addWidget(InputRow("Parking / Fines", self.parking_fines))
        vg_layout.addWidget(InputRow("Repairs / Maintenance", self.repairs))

        # --- Other Expenses Group ---
        other_group = QGroupBox("OTHER EXPENSES (WEEKLY)")
        og_layout = QVBoxLayout(other_group)
        og_layout.setSpacing(6)

        self.phone_data = MoneySpinBox(max_val=99.99)
        self.uniform_equipment = MoneySpinBox(max_val=99.99)
        self.other_costs = MoneySpinBox(max_val=999.99)

        og_layout.addWidget(InputRow("Phone / Data", self.phone_data))
        og_layout.addWidget(InputRow("Uniform / Equipment", self.uniform_equipment))
        og_layout.addWidget(InputRow("Other Costs", self.other_costs))

        # --- Tax Group ---
        tax_group = QGroupBox("TAX & EMPLOYMENT")
        tg_layout = QVBoxLayout(tax_group)
        tg_layout.setSpacing(6)

        self.employment_type = QComboBox()
        self.employment_type.addItems(["Self Employed", "PAYE Employee"])
        self.employment_type.setFixedHeight(34)

        tg_note = QLabel(
            "Tax calculated using 2025/26 HMRC rates. "
            "PAYE: Tax on gross. Self-Employed: Tax on net profit."
        )
        tg_note.setStyleSheet("color: #5a5e72; font-size: 11px; font-style: italic;")
        tg_note.setWordWrap(True)

        tg_layout.addWidget(InputRow("Employment Type", self.employment_type))
        tg_layout.addWidget(tg_note)

        # --- Calculate Button ---
        calc_btn = QPushButton("  ⚡  CALCULATE PROFIT")
        calc_btn.setFixedHeight(44)
        calc_btn.setStyleSheet("""
            QPushButton {
                background-color: #1e6fff;
                color: white;
                font-size: 14px;
                font-weight: 800;
                border-radius: 8px;
                letter-spacing: 0.8px;
            }
            QPushButton:hover { background-color: #3d7fff; }
            QPushButton:pressed { background-color: #1558cc; }
        """)
        calc_btn.clicked.connect(self.calculation_requested)

        main_layout.addWidget(income_group)
        main_layout.addWidget(vehicle_group)
        main_layout.addWidget(other_group)
        main_layout.addWidget(tax_group)
        main_layout.addWidget(calc_btn)
        main_layout.addStretch()

    def get_profile_data(self) -> dict:
        return {
            "day_rate": self.day_rate.value(),
            "days_per_week": self.days_per_week.value(),
            "hours_per_day": self.hours_per_day.value(),
            "mileage_bonus_weekly": self.mileage_bonus.value(),
            "performance_bonus_weekly": self.performance_bonus.value(),
            "training_payment": self.training_payment.value(),
            "van_rental": self.van_rental.value(),
            "fuel": self.fuel.value(),
            "insurance": self.insurance.value(),
            "parking_fines": self.parking_fines.value(),
            "repairs": self.repairs.value(),
            "phone_data": self.phone_data.value(),
            "uniform_equipment": self.uniform_equipment.value(),
            "other_costs": self.other_costs.value(),
            "employment_type": self.employment_type.currentText(),
        }

    def load_profile_data(self, data: dict):
        spinboxes = {
            "day_rate": self.day_rate,
            "mileage_bonus_weekly": self.mileage_bonus,
            "performance_bonus_weekly": self.performance_bonus,
            "training_payment": self.training_payment,
            "van_rental": self.van_rental,
            "fuel": self.fuel,
            "insurance": self.insurance,
            "parking_fines": self.parking_fines,
            "repairs": self.repairs,
            "phone_data": self.phone_data,
            "uniform_equipment": self.uniform_equipment,
            "other_costs": self.other_costs,
        }
        for key, widget in spinboxes.items():
            if key in data:
                widget.setValue(float(data[key]))

        if "days_per_week" in data:
            self.days_per_week.setValue(int(data["days_per_week"]))
        if "hours_per_day" in data:
            self.hours_per_day.setValue(float(data["hours_per_day"]))
        if "employment_type" in data:
            idx = self.employment_type.findText(data["employment_type"])
            if idx >= 0:
                self.employment_type.setCurrentIndex(idx)


# ---------------------------------------------------------------------------
# Tab: Results
# ---------------------------------------------------------------------------

class ResultsTab(QScrollArea):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWidgetResizable(True)

        container = QWidget()
        self.setWidget(container)
        self.main_layout = QVBoxLayout(container)
        self.main_layout.setSpacing(14)
        self.main_layout.setContentsMargins(16, 16, 16, 16)

        # Warnings area
        self.warnings_container = QWidget()
        self.warnings_layout = QVBoxLayout(self.warnings_container)
        self.warnings_layout.setSpacing(6)
        self.warnings_layout.setContentsMargins(0, 0, 0, 0)

        # Metric cards - Daily
        self._add_section("DAILY BREAKDOWN")
        daily_row = QHBoxLayout()
        self.card_gross_daily = MetricCard("Gross Daily", "—", "#1e6fff")
        self.card_net_daily = MetricCard("Net Daily", "—", "#27ae60")
        self.card_hourly_net = MetricCard("Net Hourly Rate", "—", "#9b59b6")
        daily_row.addWidget(self.card_gross_daily)
        daily_row.addWidget(self.card_net_daily)
        daily_row.addWidget(self.card_hourly_net)

        # Weekly
        self._add_section("WEEKLY BREAKDOWN")
        weekly_row = QHBoxLayout()
        self.card_gross_weekly = MetricCard("Gross Weekly", "—", "#1e6fff")
        self.card_expenses_weekly = MetricCard("Weekly Expenses", "—", "#e74c3c")
        self.card_tax_weekly = MetricCard("Tax & NI (weekly)", "—", "#f39c12")
        self.card_net_weekly = MetricCard("Net Weekly", "—", "#27ae60")
        weekly_row.addWidget(self.card_gross_weekly)
        weekly_row.addWidget(self.card_expenses_weekly)
        weekly_row.addWidget(self.card_tax_weekly)
        weekly_row.addWidget(self.card_net_weekly)

        # Monthly
        self._add_section("MONTHLY BREAKDOWN")
        monthly_row = QHBoxLayout()
        self.card_gross_monthly = MetricCard("Gross Monthly", "—", "#1e6fff")
        self.card_expenses_monthly = MetricCard("Monthly Expenses", "—", "#e74c3c")
        self.card_tax_monthly = MetricCard("Tax & NI (monthly)", "—", "#f39c12")
        self.card_net_monthly = MetricCard("Net Monthly", "—", "#27ae60")
        monthly_row.addWidget(self.card_gross_monthly)
        monthly_row.addWidget(self.card_expenses_monthly)
        monthly_row.addWidget(self.card_tax_monthly)
        monthly_row.addWidget(self.card_net_monthly)

        # Annual
        self._add_section("ANNUAL PROJECTION")
        annual_row = QHBoxLayout()
        self.card_gross_annual = MetricCard("Gross Annual", "—", "#1e6fff")
        self.card_expenses_annual = MetricCard("Annual Expenses", "—", "#e74c3c")
        self.card_tax_annual = MetricCard("Total Tax & NI", "—", "#f39c12")
        self.card_net_annual = MetricCard("Take-Home Annual", "—", "#27ae60")
        annual_row.addWidget(self.card_gross_annual)
        annual_row.addWidget(self.card_expenses_annual)
        annual_row.addWidget(self.card_tax_annual)
        annual_row.addWidget(self.card_net_annual)

        # Tax Detail
        self.tax_detail_box = QGroupBox("TAX DETAIL")
        tax_detail_layout = QGridLayout(self.tax_detail_box)
        self.lbl_income_tax = QLabel("£ —")
        self.lbl_ni = QLabel("£ —")
        self.lbl_tax_pct = QLabel("—%")
        self.lbl_employment_type = QLabel("—")
        for lbl in [self.lbl_income_tax, self.lbl_ni, self.lbl_tax_pct, self.lbl_employment_type]:
            lbl.setStyleSheet("color: #e8eaf0; font-weight: 600; font-size: 13px;")

        tax_detail_layout.addWidget(QLabel("Employment Type:"), 0, 0)
        tax_detail_layout.addWidget(self.lbl_employment_type, 0, 1)
        tax_detail_layout.addWidget(QLabel("Income Tax (annual):"), 1, 0)
        tax_detail_layout.addWidget(self.lbl_income_tax, 1, 1)
        tax_detail_layout.addWidget(QLabel("National Insurance (annual):"), 2, 0)
        tax_detail_layout.addWidget(self.lbl_ni, 2, 1)
        tax_detail_layout.addWidget(QLabel("Tax Reserve:"), 3, 0)
        tax_detail_layout.addWidget(self.lbl_tax_pct, 3, 1)
        for i in range(2):
            tax_detail_layout.setColumnStretch(i, 1)

        # Assemble
        self.main_layout.addWidget(self.warnings_container)
        self.main_layout.addLayout(daily_row)
        self.main_layout.addLayout(weekly_row)
        self.main_layout.addLayout(monthly_row)
        self.main_layout.addLayout(annual_row)
        self.main_layout.addWidget(self.tax_detail_box)
        self.main_layout.addStretch()

        self._section_widgets = []

    def _add_section(self, title: str):
        lbl = SectionHeader(title)
        self.main_layout.addWidget(lbl)

    def update_results(self, r: dict):
        # Clear warnings
        for i in reversed(range(self.warnings_layout.count())):
            w = self.warnings_layout.itemAt(i).widget()
            if w:
                w.deleteLater()

        for msg in r.get("warnings", []):
            lbl = QLabel(msg)
            lbl.setWordWrap(True)
            lbl.setStyleSheet("""
                color: #f39c12;
                font-size: 12px;
                font-weight: 600;
                padding: 8px 12px;
                background-color: #2d2408;
                border-left: 3px solid #f39c12;
                border-radius: 4px;
            """)
            self.warnings_layout.addWidget(lbl)

        def fmt(v): return f"£{v:,.2f}"

        self.card_gross_daily.update_value(fmt(r["gross_daily"]))
        net_color = "#27ae60" if r["net_daily"] >= 0 else "#e74c3c"
        self.card_net_daily.set_color(net_color)
        self.card_net_daily.update_value(fmt(r["net_daily"]))

        hw_color = "#27ae60" if r["hourly_rate_net"] >= UK_TAX["minimum_wage"] else "#e74c3c"
        self.card_hourly_net.set_color(hw_color)
        self.card_hourly_net.update_value(fmt(r["hourly_rate_net"]) + "/hr")

        self.card_gross_weekly.update_value(fmt(r["gross_weekly"]))
        self.card_expenses_weekly.update_value(fmt(r["weekly_expenses"]))
        self.card_tax_weekly.update_value(fmt(r["tax_weekly"]))
        nw_color = "#27ae60" if r["net_weekly"] >= 0 else "#e74c3c"
        self.card_net_weekly.set_color(nw_color)
        self.card_net_weekly.update_value(fmt(r["net_weekly"]))

        self.card_gross_monthly.update_value(fmt(r["gross_monthly"]))
        self.card_expenses_monthly.update_value(fmt(r["monthly_expenses"]))
        self.card_tax_monthly.update_value(fmt(r["tax_monthly"]))
        nm_color = "#27ae60" if r["net_monthly"] >= 0 else "#e74c3c"
        self.card_net_monthly.set_color(nm_color)
        self.card_net_monthly.update_value(fmt(r["net_monthly"]))

        self.card_gross_annual.update_value(fmt(r["gross_annual"]))
        self.card_expenses_annual.update_value(fmt(r["annual_expenses"]))
        self.card_tax_annual.update_value(fmt(r["total_tax_annual"]))
        na_color = "#27ae60" if r["net_annual"] >= 0 else "#e74c3c"
        self.card_net_annual.set_color(na_color)
        self.card_net_annual.update_value(fmt(r["net_annual"]))

        self.lbl_employment_type.setText(r["employment_type"])
        self.lbl_income_tax.setText(fmt(r["income_tax_annual"]))
        self.lbl_ni.setText(fmt(r["ni_annual"]))
        self.lbl_tax_pct.setText(f"{r['tax_reserve_pct']:.1f}% of gross")


# ---------------------------------------------------------------------------
# Tab: Charts
# ---------------------------------------------------------------------------

class ChartsTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(14)

        if not HAS_MATPLOTLIB:
            lbl = QLabel("📊 Charts require matplotlib.\nInstall with: pip install matplotlib")
            lbl.setAlignment(Qt.AlignCenter)
            lbl.setStyleSheet("color: #7b8094; font-size: 14px;")
            layout.addWidget(lbl)
            self.has_charts = False
            return

        self.has_charts = True

        # Create figure with 3 subplots
        plt.style.use('dark_background')
        self.figure = Figure(figsize=(12, 9), facecolor='#13161f')
        self.canvas = FigureCanvas(self.figure)
        self.canvas.setStyleSheet("background-color: #13161f; border-radius: 8px;")

        layout.addWidget(self.canvas)
        self._last_results = None

    def update_charts(self, results: dict):
        if not self.has_charts:
            return
        self._last_results = results
        self.figure.clear()

        r = results
        dark_bg = '#13161f'
        grid_color = '#2a2d3a'
        text_color = '#c0c4d4'

        # --- Chart 1: Income vs Expenses (Weekly) ---
        ax1 = self.figure.add_subplot(2, 2, 1)
        ax1.set_facecolor('#1a1d28')
        categories = ['Gross\nIncome', 'Vehicle\nCosts', 'Other\nCosts', 'Tax &\nNI', 'Net\nProfit']
        values = [
            r['gross_weekly'],
            r['vehicle_costs_weekly'],
            r['other_costs_weekly'],
            r['tax_weekly'],
            max(0, r['net_weekly'])
        ]
        colors = ['#1e6fff', '#e74c3c', '#f39c12', '#9b59b6', '#27ae60']
        bars = ax1.bar(categories, values, color=colors, width=0.6, edgecolor='none')
        ax1.set_title('Weekly Income vs Expenses', color=text_color, fontsize=11, fontweight='bold', pad=10)
        ax1.set_ylabel('£', color=text_color, fontsize=9)
        ax1.tick_params(colors=text_color, labelsize=8)
        ax1.spines['bottom'].set_color(grid_color)
        ax1.spines['top'].set_visible(False)
        ax1.spines['right'].set_visible(False)
        ax1.spines['left'].set_color(grid_color)
        ax1.yaxis.grid(True, color=grid_color, alpha=0.5)
        ax1.set_axisbelow(True)
        for bar, val in zip(bars, values):
            ax1.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 1,
                     f'£{val:.0f}', ha='center', va='bottom', color=text_color, fontsize=8)

        # --- Chart 2: Expense Breakdown Pie ---
        ax2 = self.figure.add_subplot(2, 2, 2)
        ax2.set_facecolor('#1a1d28')
        exp_labels = ['Van Rental', 'Fuel', 'Insurance', 'Parking/Fines', 'Repairs', 'Phone/Data', 'Uniform', 'Other']
        exp_values_raw = [
            r.get('van_rental_w', 0), r.get('fuel_w', 0), r.get('insurance_w', 0),
            r.get('parking_w', 0), r.get('repairs_w', 0), r.get('phone_w', 0),
            r.get('uniform_w', 0), r.get('other_w', 0)
        ]
        # Use total expense data we have
        total_exp = r['weekly_expenses']
        if total_exp > 0:
            pie_labels = ['Vehicle', 'Other', 'Tax & NI']
            pie_values = [r['vehicle_costs_weekly'], r['other_costs_weekly'], r['tax_weekly']]
            pie_colors = ['#e74c3c', '#f39c12', '#9b59b6']
            wedges, texts, autotexts = ax2.pie(
                pie_values, labels=pie_labels, colors=pie_colors,
                autopct='%1.1f%%', startangle=90,
                textprops={'color': text_color, 'fontsize': 9},
                wedgeprops={'edgecolor': '#13161f', 'linewidth': 2}
            )
            for autotext in autotexts:
                autotext.set_color(text_color)
                autotext.set_fontsize(8)
        else:
            ax2.text(0.5, 0.5, 'No expenses', ha='center', va='center', color=text_color)
        ax2.set_title('Weekly Cost Breakdown', color=text_color, fontsize=11, fontweight='bold', pad=10)

        # --- Chart 3: Monthly Projection (12 months) ---
        ax3 = self.figure.add_subplot(2, 1, 2)
        ax3.set_facecolor('#1a1d28')
        months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        monthly_gross = [r['gross_monthly']] * 12
        monthly_net = [r['net_monthly']] * 12

        x = range(12)
        ax3.fill_between(x, monthly_gross, alpha=0.25, color='#1e6fff')
        ax3.fill_between(x, monthly_net, alpha=0.35, color='#27ae60')
        ax3.plot(x, monthly_gross, color='#1e6fff', linewidth=2.5, label='Gross Monthly', marker='o', markersize=4)
        ax3.plot(x, monthly_net, color='#27ae60', linewidth=2.5, label='Net Monthly', marker='o', markersize=4)
        ax3.axhline(y=0, color='#e74c3c', linestyle='--', alpha=0.5, linewidth=1)
        ax3.set_xticks(x)
        ax3.set_xticklabels(months, color=text_color, fontsize=9)
        ax3.set_title('Annual Monthly Projection', color=text_color, fontsize=11, fontweight='bold', pad=10)
        ax3.set_ylabel('£', color=text_color, fontsize=9)
        ax3.tick_params(colors=text_color, labelsize=8)
        ax3.spines['bottom'].set_color(grid_color)
        ax3.spines['top'].set_visible(False)
        ax3.spines['right'].set_visible(False)
        ax3.spines['left'].set_color(grid_color)
        ax3.yaxis.grid(True, color=grid_color, alpha=0.4)
        ax3.set_axisbelow(True)
        legend = ax3.legend(facecolor='#1a1d28', edgecolor=grid_color, labelcolor=text_color, fontsize=9)

        self.figure.patch.set_facecolor(dark_bg)
        self.figure.tight_layout(pad=2.5)
        self.canvas.draw()


# ---------------------------------------------------------------------------
# Tab: Scenario Comparison
# ---------------------------------------------------------------------------

class ScenarioTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(14)

        header = SectionHeader("Scenario Comparison")
        layout.addWidget(header)

        note = QLabel("Compares Good / Realistic / Worst case based on your current inputs.")
        note.setStyleSheet("color: #7b8094; font-size: 12px;")
        layout.addWidget(note)

        # Table
        self.table = QTableWidget(8, 4)
        self.table.setHorizontalHeaderLabels(["Metric", "✅ Good", "📊 Realistic", "⚠ Worst"])
        self.table.verticalHeader().setVisible(False)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet(self.table.styleSheet() + """
            QTableWidget { alternate-background-color: #161921; }
        """)
        layout.addWidget(self.table)

        # Chart
        if HAS_MATPLOTLIB:
            plt.style.use('dark_background')
            self.scenario_figure = Figure(figsize=(10, 4), facecolor='#13161f')
            self.scenario_canvas = FigureCanvas(self.scenario_figure)
            layout.addWidget(self.scenario_canvas)
        else:
            self.scenario_figure = None
            self.scenario_canvas = None

        layout.addStretch()

    def update_scenarios(self, good: dict, realistic: dict, worst: dict):
        rows = [
            ("Gross Weekly", "gross_weekly"),
            ("Weekly Expenses", "weekly_expenses"),
            ("Tax & NI (weekly)", "tax_weekly"),
            ("Net Weekly", "net_weekly"),
            ("Net Daily", "net_daily"),
            ("Net Hourly", "hourly_rate_net"),
            ("Net Monthly", "net_monthly"),
            ("Net Annual", "net_annual"),
        ]
        self.table.setRowCount(len(rows))

        scenarios = [good, realistic, worst]
        colors = ["#27ae60", "#1e6fff", "#e74c3c"]

        for row_idx, (label, key) in enumerate(rows):
            item0 = QTableWidgetItem(label)
            item0.setForeground(QColor("#c0c4d4"))
            self.table.setItem(row_idx, 0, item0)

            for col_idx, (sc, color) in enumerate(zip(scenarios, colors)):
                val = sc.get(key, 0)
                if key == "hourly_rate_net":
                    text = f"£{val:.2f}/hr"
                else:
                    text = f"£{val:,.2f}"
                item = QTableWidgetItem(text)
                item.setTextAlignment(Qt.AlignCenter)
                item.setForeground(QColor(color))
                self.table.setItem(row_idx, col_idx + 1, item)

        # Update scenario chart
        if self.scenario_figure:
            self.scenario_figure.clear()
            ax = self.scenario_figure.add_subplot(111)
            ax.set_facecolor('#1a1d28')

            labels = ['Good', 'Realistic', 'Worst']
            net_annuals = [good['net_annual'], realistic['net_annual'], worst['net_annual']]
            colors_chart = ['#27ae60', '#1e6fff', '#e74c3c']

            bars = ax.bar(labels, net_annuals, color=colors_chart, width=0.45, edgecolor='none')
            ax.axhline(0, color='#7b8094', linestyle='--', alpha=0.5)
            ax.set_title('Annual Take-Home by Scenario', color='#c0c4d4', fontsize=12, fontweight='bold')
            ax.set_ylabel('£', color='#c0c4d4')
            ax.tick_params(colors='#c0c4d4')
            for spine in ['top', 'right']:
                ax.spines[spine].set_visible(False)
            ax.spines['bottom'].set_color('#2a2d3a')
            ax.spines['left'].set_color('#2a2d3a')
            ax.yaxis.grid(True, color='#2a2d3a', alpha=0.5)
            ax.set_axisbelow(True)

            for bar, val in zip(bars, net_annuals):
                ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 50,
                        f'£{val:,.0f}', ha='center', va='bottom', color='#e8eaf0', fontsize=10, fontweight='bold')

            self.scenario_figure.patch.set_facecolor('#13161f')
            self.scenario_figure.tight_layout(pad=2)
            self.scenario_canvas.draw()


# ---------------------------------------------------------------------------
# Jobs Manager Panel (Left Sidebar)
# ---------------------------------------------------------------------------

class JobsPanel(QWidget):
    job_selected = Signal(str)  # job name
    job_added = Signal(str)
    job_deleted = Signal(str)

    PRESETS = [
        "Amazon DSP",
        "DPD",
        "Evri",
        "Uber Delivery",
        "Self-employed Van",
        "Royal Mail",
        "Yodel",
        "Hermes",
    ]

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFixedWidth(210)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(8)

        # Header
        header = QLabel("SAVED JOBS")
        header.setStyleSheet("""
            color: #7b8094;
            font-size: 10px;
            font-weight: 800;
            letter-spacing: 1.5px;
            padding: 4px 6px;
        """)
        layout.addWidget(header)

        self.job_list = QListWidget()
        self.job_list.itemClicked.connect(lambda item: self.job_selected.emit(item.text()))
        layout.addWidget(self.job_list)

        # Add from presets
        preset_label = QLabel("ADD PRESET:")
        preset_label.setStyleSheet("color: #5a5e72; font-size: 10px; font-weight: 700; letter-spacing: 1px;")
        self.preset_combo = QComboBox()
        self.preset_combo.addItems(self.PRESETS)
        self.preset_combo.setFixedHeight(30)

        add_preset_btn = QPushButton("Add Preset")
        add_preset_btn.setFixedHeight(30)
        add_preset_btn.setStyleSheet("font-size: 11px; padding: 4px 8px;")
        add_preset_btn.clicked.connect(self._add_preset)

        # Custom name
        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("Custom job name...")
        self.name_input.setFixedHeight(30)

        add_custom_btn = QPushButton("Add Custom")
        add_custom_btn.setFixedHeight(30)
        add_custom_btn.setStyleSheet("font-size: 11px; padding: 4px 8px;")
        add_custom_btn.clicked.connect(self._add_custom)

        del_btn = QPushButton("Delete Selected")
        del_btn.setFixedHeight(30)
        del_btn.setStyleSheet("""
            QPushButton { background-color: #c0392b; font-size: 11px; padding: 4px 8px; }
            QPushButton:hover { background-color: #e74c3c; }
        """)
        del_btn.clicked.connect(self._delete_selected)

        layout.addWidget(preset_label)
        layout.addWidget(self.preset_combo)
        layout.addWidget(add_preset_btn)
        layout.addWidget(self.name_input)
        layout.addWidget(add_custom_btn)
        layout.addWidget(del_btn)

    def _add_preset(self):
        name = self.preset_combo.currentText()
        self._add_job(name)

    def _add_custom(self):
        name = self.name_input.text().strip()
        if not name:
            return
        self._add_job(name)
        self.name_input.clear()

    def _add_job(self, name: str):
        # Prevent duplicates
        existing = [self.job_list.item(i).text() for i in range(self.job_list.count())]
        if name in existing:
            count = sum(1 for n in existing if n.startswith(name))
            name = f"{name} ({count})"
        self.job_list.addItem(QListWidgetItem(name))
        self.job_added.emit(name)

    def _delete_selected(self):
        items = self.job_list.selectedItems()
        for item in items:
            name = item.text()
            self.job_list.takeItem(self.job_list.row(item))
            self.job_deleted.emit(name)

    def get_all_jobs(self) -> list:
        return [self.job_list.item(i).text() for i in range(self.job_list.count())]

    def select_job(self, name: str):
        for i in range(self.job_list.count()):
            if self.job_list.item(i).text() == name:
                self.job_list.setCurrentRow(i)
                break


# ---------------------------------------------------------------------------
# Export Functions
# ---------------------------------------------------------------------------

def export_to_csv(results: dict, profile_name: str, filepath: str):
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(["Delivery Driver Profit Calculator UK 2026"])
        writer.writerow(["Profile:", profile_name])
        writer.writerow(["Generated:", datetime.now().strftime("%d/%m/%Y %H:%M")])
        writer.writerow([])
        writer.writerow(["Metric", "Value"])
        rows = [
            ("Gross Daily", f"£{results['gross_daily']:.2f}"),
            ("Net Daily", f"£{results['net_daily']:.2f}"),
            ("Net Hourly Rate", f"£{results['hourly_rate_net']:.2f}/hr"),
            ("Gross Weekly", f"£{results['gross_weekly']:.2f}"),
            ("Weekly Expenses", f"£{results['weekly_expenses']:.2f}"),
            ("Tax & NI (weekly)", f"£{results['tax_weekly']:.2f}"),
            ("Net Weekly", f"£{results['net_weekly']:.2f}"),
            ("Gross Monthly", f"£{results['gross_monthly']:.2f}"),
            ("Monthly Expenses", f"£{results['monthly_expenses']:.2f}"),
            ("Tax & NI (monthly)", f"£{results['tax_monthly']:.2f}"),
            ("Net Monthly", f"£{results['net_monthly']:.2f}"),
            ("Gross Annual", f"£{results['gross_annual']:.2f}"),
            ("Annual Expenses", f"£{results['annual_expenses']:.2f}"),
            ("Income Tax (annual)", f"£{results['income_tax_annual']:.2f}"),
            ("NI (annual)", f"£{results['ni_annual']:.2f}"),
            ("Net Annual Take-Home", f"£{results['net_annual']:.2f}"),
            ("Tax Reserve %", f"{results['tax_reserve_pct']:.1f}%"),
            ("Employment Type", results['employment_type']),
        ]
        for row in rows:
            writer.writerow(row)
        if results.get("warnings"):
            writer.writerow([])
            writer.writerow(["WARNINGS"])
            for w in results["warnings"]:
                writer.writerow([w])


def export_to_excel(results: dict, profile_name: str, filepath: str):
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl not installed")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Profit Report"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill("solid", fgColor="1E6FFF")
    title_font = Font(bold=True, color="FFFFFF", size=14)
    title_fill = PatternFill("solid", fgColor="0F1117")
    section_font = Font(bold=True, color="1E6FFF", size=11)
    positive_font = Font(color="27AE60", bold=True)
    negative_font = Font(color="E74C3C", bold=True)
    warn_font = Font(color="F39C12", bold=True)

    def write_header(row, text, cols=2):
        ws.cell(row=row, column=1, value=text).font = section_font
        return row + 1

    def write_row(row, label, value, is_total=False):
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=2, value=value)
        c1.alignment = Alignment(horizontal='left')
        c2.alignment = Alignment(horizontal='right')
        if is_total:
            c1.font = Font(bold=True)
            c2.font = positive_font if (isinstance(value, str) and not value.startswith('-')) else Font(bold=True)
        return row + 1

    r = 1
    ws.cell(r, 1, "Delivery Driver Profit Calculator UK 2026").font = title_font
    ws.cell(r, 1).fill = title_fill
    r += 1
    ws.cell(r, 1, f"Profile: {profile_name}")
    r += 1
    ws.cell(r, 1, f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    r += 2

    r = write_header(r, "DAILY")
    r = write_row(r, "Gross Daily", f"£{results['gross_daily']:.2f}")
    r = write_row(r, "Net Daily", f"£{results['net_daily']:.2f}", True)
    r = write_row(r, "Net Hourly Rate", f"£{results['hourly_rate_net']:.2f}/hr")
    r += 1

    r = write_header(r, "WEEKLY")
    r = write_row(r, "Gross Weekly", f"£{results['gross_weekly']:.2f}")
    r = write_row(r, "Weekly Expenses", f"£{results['weekly_expenses']:.2f}")
    r = write_row(r, "Tax & NI (weekly)", f"£{results['tax_weekly']:.2f}")
    r = write_row(r, "Net Weekly Profit", f"£{results['net_weekly']:.2f}", True)
    r += 1

    r = write_header(r, "MONTHLY")
    r = write_row(r, "Gross Monthly", f"£{results['gross_monthly']:.2f}")
    r = write_row(r, "Monthly Expenses", f"£{results['monthly_expenses']:.2f}")
    r = write_row(r, "Tax & NI (monthly)", f"£{results['tax_monthly']:.2f}")
    r = write_row(r, "Net Monthly", f"£{results['net_monthly']:.2f}", True)
    r += 1

    r = write_header(r, "ANNUAL")
    r = write_row(r, "Gross Annual", f"£{results['gross_annual']:.2f}")
    r = write_row(r, "Annual Expenses", f"£{results['annual_expenses']:.2f}")
    r = write_row(r, "Income Tax", f"£{results['income_tax_annual']:.2f}")
    r = write_row(r, "National Insurance", f"£{results['ni_annual']:.2f}")
    r = write_row(r, "Net Annual Take-Home", f"£{results['net_annual']:.2f}", True)
    r += 1

    if results.get("warnings"):
        r = write_header(r, "WARNINGS")
        for w in results["warnings"]:
            ws.cell(r, 1, w).font = warn_font
            r += 1

    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 20
    wb.save(filepath)


# ---------------------------------------------------------------------------
# Main Window
# ---------------------------------------------------------------------------

class MainWindow(QMainWindow):
    SAVE_FILE = Path.home() / ".delivery_calculator_jobs.json"

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Delivery Driver Profit Calculator UK 2026")
        self.setMinimumSize(1100, 720)
        self.resize(1280, 800)
        self.setStyleSheet(STYLE_SHEET)

        # Data
        self.profiles: dict[str, JobProfile] = {}
        self.current_profile_name: str | None = None
        self.last_results: dict = {}

        self._build_ui()
        self._load_saved_jobs()

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        root_layout = QHBoxLayout(central)
        root_layout.setContentsMargins(0, 0, 0, 0)
        root_layout.setSpacing(0)

        # Sidebar
        sidebar_container = QWidget()
        sidebar_container.setFixedWidth(220)
        sidebar_container.setStyleSheet("background-color: #0c0e15; border-right: 1px solid #1e2130;")
        sidebar_layout = QVBoxLayout(sidebar_container)
        sidebar_layout.setContentsMargins(0, 0, 0, 0)

        # Logo/App name
        logo_widget = QWidget()
        logo_widget.setStyleSheet("background-color: #0c0e15; padding: 16px 12px;")
        logo_layout = QVBoxLayout(logo_widget)
        logo_layout.setSpacing(2)
        logo_layout.setContentsMargins(12, 12, 12, 12)
        app_name = QLabel("🚐 PROFIT CALC")
        app_name.setStyleSheet("color: #1e6fff; font-size: 13px; font-weight: 900; letter-spacing: 1px;")
        app_subtitle = QLabel("UK Delivery Driver 2026")
        app_subtitle.setStyleSheet("color: #3a3d52; font-size: 10px; font-weight: 600;")
        logo_layout.addWidget(app_name)
        logo_layout.addWidget(app_subtitle)

        self.jobs_panel = JobsPanel()
        self.jobs_panel.job_selected.connect(self._on_job_selected)
        self.jobs_panel.job_added.connect(self._on_job_added)
        self.jobs_panel.job_deleted.connect(self._on_job_deleted)

        sidebar_layout.addWidget(logo_widget)
        sidebar_layout.addWidget(self.jobs_panel)

        # Main content area
        content = QWidget()
        content_layout = QVBoxLayout(content)
        content_layout.setContentsMargins(0, 0, 0, 0)
        content_layout.setSpacing(0)

        # Toolbar
        toolbar = self._build_toolbar()
        content_layout.addWidget(toolbar)

        # Tabs
        self.tabs = QTabWidget()
        self.tabs.setDocumentMode(True)

        self.inputs_tab = InputsTab()
        self.inputs_tab.calculation_requested.connect(self.run_calculation)

        self.results_tab = ResultsTab()
        self.charts_tab = ChartsTab()
        self.scenario_tab = ScenarioTab()

        self.tabs.addTab(self.inputs_tab, "  Inputs  ")
        self.tabs.addTab(self.results_tab, "  Results  ")
        self.tabs.addTab(self.charts_tab, "  Charts  ")
        self.tabs.addTab(self.scenario_tab, "  Scenarios  ")

        content_layout.addWidget(self.tabs)

        # Status bar
        self.status_bar = QLabel("Ready. Select or create a job profile to begin.")
        self.status_bar.setStyleSheet("""
            background-color: #0c0e15;
            color: #5a5e72;
            font-size: 11px;
            padding: 6px 16px;
            border-top: 1px solid #1e2130;
        """)
        content_layout.addWidget(self.status_bar)

        root_layout.addWidget(sidebar_container)
        root_layout.addWidget(content)

    def _build_toolbar(self) -> QWidget:
        toolbar = QWidget()
        toolbar.setFixedHeight(50)
        toolbar.setStyleSheet("background-color: #0f1117; border-bottom: 1px solid #1e2130;")
        layout = QHBoxLayout(toolbar)
        layout.setContentsMargins(14, 6, 14, 6)
        layout.setSpacing(8)

        self.profile_label = QLabel("No profile selected")
        self.profile_label.setStyleSheet("color: #c0c4d4; font-weight: 700; font-size: 13px;")
        layout.addWidget(self.profile_label)
        layout.addStretch()

        save_btn = QPushButton("💾 Save Profile")
        save_btn.setFixedHeight(34)
        save_btn.setStyleSheet("""
            QPushButton { background-color: #252836; color: #c0c4d4; font-size: 11px; padding: 4px 14px; }
            QPushButton:hover { background-color: #2e3245; }
        """)
        save_btn.clicked.connect(self._save_current_profile)

        calc_btn = QPushButton("⚡ Calculate")
        calc_btn.setFixedHeight(34)
        calc_btn.clicked.connect(self.run_calculation)

        export_csv_btn = QPushButton("📄 CSV")
        export_csv_btn.setFixedHeight(34)
        export_csv_btn.setStyleSheet("""
            QPushButton { background-color: #27ae60; color: white; font-size: 11px; padding: 4px 14px; }
            QPushButton:hover { background-color: #2ecc71; }
        """)
        export_csv_btn.clicked.connect(self._export_csv)

        export_xlsx_btn = QPushButton("📊 Excel")
        export_xlsx_btn.setFixedHeight(34)
        export_xlsx_btn.setStyleSheet("""
            QPushButton { background-color: #1e6fff; color: white; font-size: 11px; padding: 4px 14px; }
            QPushButton:hover { background-color: #3d7fff; }
        """)
        export_xlsx_btn.clicked.connect(self._export_excel)

        layout.addWidget(save_btn)
        layout.addWidget(calc_btn)
        layout.addWidget(export_csv_btn)
        layout.addWidget(export_xlsx_btn)
        return toolbar

    def _on_job_added(self, name: str):
        profile = JobProfile(name)
        self.profiles[name] = profile
        self.current_profile_name = name
        self.profile_label.setText(f"Profile: {name}")
        self.jobs_panel.select_job(name)
        self.status_bar.setText(f"New profile '{name}' created. Configure inputs and calculate.")

    def _on_job_selected(self, name: str):
        # Save current before switching
        if self.current_profile_name and self.current_profile_name in self.profiles:
            self._update_profile_from_inputs(self.current_profile_name)

        self.current_profile_name = name
        self.profile_label.setText(f"Profile: {name}")
        profile = self.profiles.get(name)
        if profile:
            self.inputs_tab.load_profile_data(profile.to_dict())
        self.status_bar.setText(f"Profile '{name}' loaded. Press Calculate to update results.")

    def _on_job_deleted(self, name: str):
        self.profiles.pop(name, None)
        if self.current_profile_name == name:
            self.current_profile_name = None
            self.profile_label.setText("No profile selected")
        self._persist_jobs()

    def _update_profile_from_inputs(self, name: str):
        if name not in self.profiles:
            return
        data = self.inputs_tab.get_profile_data()
        profile = self.profiles[name]
        for k, v in data.items():
            if hasattr(profile, k):
                setattr(profile, k, v)

    def _save_current_profile(self):
        if not self.current_profile_name:
            QMessageBox.information(self, "No Profile", "Please add or select a job profile first.")
            return
        self._update_profile_from_inputs(self.current_profile_name)
        self._persist_jobs()
        self.status_bar.setText(f"Profile '{self.current_profile_name}' saved.")

    def run_calculation(self):
        if not self.current_profile_name:
            # Create a temp profile
            name = "Quick Calculation"
            self.profiles[name] = JobProfile(name)
            self.current_profile_name = name
            self.profile_label.setText(f"Profile: {name}")

        self._update_profile_from_inputs(self.current_profile_name)
        profile = self.profiles[self.current_profile_name]

        calc = ProfitCalculator(profile)
        results = calc.calculate()
        self.last_results = results

        # Update all tabs
        self.results_tab.update_results(results)
        self.charts_tab.update_charts(results)

        # Scenarios
        good = calc.get_scenario("Good")
        realistic = calc.get_scenario("Realistic")
        worst = calc.get_scenario("Worst")
        self.scenario_tab.update_scenarios(good, realistic, worst)

        warning_count = len(results.get("warnings", []))
        warn_text = f" | ⚠ {warning_count} warning(s)" if warning_count else ""
        self.status_bar.setText(
            f"Calculated: Net Weekly £{results['net_weekly']:.2f} | "
            f"Net Annual £{results['net_annual']:.2f} | "
            f"Tax Rate: {results['tax_reserve_pct']:.1f}%{warn_text}"
        )
        # Switch to results
        self.tabs.setCurrentIndex(1)

    def _export_csv(self):
        if not self.last_results:
            QMessageBox.information(self, "No Data", "Please calculate first.")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Export CSV", f"delivery_profit_{self.current_profile_name or 'report'}.csv",
            "CSV Files (*.csv)"
        )
        if path:
            try:
                export_to_csv(self.last_results, self.current_profile_name or "—", path)
                self.status_bar.setText(f"Exported CSV: {path}")
            except Exception as e:
                QMessageBox.critical(self, "Export Error", str(e))

    def _export_excel(self):
        if not self.last_results:
            QMessageBox.information(self, "No Data", "Please calculate first.")
            return
        if not HAS_OPENPYXL:
            QMessageBox.warning(self, "Missing Library", "Install openpyxl: pip install openpyxl")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Excel", f"delivery_profit_{self.current_profile_name or 'report'}.xlsx",
            "Excel Files (*.xlsx)"
        )
        if path:
            try:
                export_to_excel(self.last_results, self.current_profile_name or "—", path)
                self.status_bar.setText(f"Exported Excel: {path}")
            except Exception as e:
                QMessageBox.critical(self, "Export Error", str(e))

    def _persist_jobs(self):
        try:
            data = {name: p.to_dict() for name, p in self.profiles.items()}
            with open(self.SAVE_FILE, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2)
        except Exception as e:
            print(f"Save error: {e}")

    def _load_saved_jobs(self):
        if not self.SAVE_FILE.exists():
            return
        try:
            with open(self.SAVE_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
            for name, pdata in data.items():
                profile = JobProfile.from_dict(pdata)
                self.profiles[name] = profile
                self.jobs_panel._add_job.__func__(self.jobs_panel, name)
            # Workaround: directly add items to list since _add_job also fires signal
            # Re-populate cleanly
            self.jobs_panel.job_list.clear()
            for name in self.profiles:
                self.jobs_panel.job_list.addItem(QListWidgetItem(name))
            self.status_bar.setText(f"Loaded {len(self.profiles)} saved profile(s).")
        except Exception as e:
            print(f"Load error: {e}")

    def closeEvent(self, event):
        if self.current_profile_name:
            self._update_profile_from_inputs(self.current_profile_name)
        self._persist_jobs()
        super().closeEvent(event)


# ---------------------------------------------------------------------------
# Entry Point
# ---------------------------------------------------------------------------

def main():
    app = QApplication(sys.argv)
    app.setApplicationName("Delivery Driver Profit Calculator UK 2026")
    app.setStyle("Fusion")

    # High DPI support
    app.setAttribute(Qt.AA_UseHighDpiPixmaps, True)

    window = MainWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
